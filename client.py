#this program is designed to send 100 motor sensor count data

import socket
import time 
import openpyxl


TIMEOUT = 5

path = "sample.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

max_row = sheet.max_row


HOST = "127.0.0.1"  # The server's hostname or IP address
PORT = 12345  # The port used by the server

TIMEOUT = 2

with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as server:
    server.connect((HOST, PORT))
    for i in range(2,101):
        time.sleep(1)
        a = time.time()
        date = sheet.cell(row = i, column = 1)
        motor_count = sheet.cell(row = i, column = 2)
        message = f"Date: {date.value},  Sensor data: {motor_count.value}"
        print(f"sending {message} to server...")
        server.send(message.encode())
        data = server.recv(1024)
        b = time.time()
        time_interval = b-a
        if TIMEOUT < time_interval:
            print("the packet was lost")
        else:
            print(f"Status: {data} , Time elapsed : {time_interval}")
     
#print(f"Received {data!r}")

